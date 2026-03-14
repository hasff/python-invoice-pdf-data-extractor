from pathlib import Path
import pdfplumber
from draw_lines import draw_vertical_lines, draw_boxes
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, NamedStyle

INPUT_FILE = "Invoice_sample.pdf"
OUTPUT_FILE = "example_output.xlsx"

SPACE_BETWEEN_WORDS = 4

def __get_cell_text(cell):
    return ' '.join([word['text'] for word in cell])

def __words_to_lines(words):
    yPos = 0
    textLines = []
    textLine = []
    for w in words:
        if w['bottom'] != yPos:
            if textLine != []:
                textLines.append(textLine)
            yPos = w['bottom']
            textLine = []

        textLine.append(w) 
    textLines.append(textLine)

    return textLines

def __group_line_items_as_cells(lines, words_gap= SPACE_BETWEEN_WORDS):
    tableData = []
    for line in lines:
        tableRow = [[line[0]]]
        for item in line[1:]:
            currStart = item['x0']
            prevEnd = tableRow[-1][-1]['x1']

            # words are close enough to be in the same table cell?
            if currStart - prevEnd < words_gap: 
                tableRow[-1].append(item)
            else: 
                tableRow.append([item])

        tableData.append(tableRow)  
    return tableData  


def _extract_doc_header(doc_header_area):
    words = doc_header_area.extract_words()
    lines = __words_to_lines(words)
    headerData = __group_line_items_as_cells(lines, SPACE_BETWEEN_WORDS + 10)

    line = -1       
    cell = 0
    def next_line():
        nonlocal line
        line += 1
        return line
    company_name = __get_cell_text(headerData[next_line()][cell])
    invoice_no = __get_cell_text(headerData[next_line()][cell])
    date = __get_cell_text(headerData[next_line()][cell])
    
    return company_name, invoice_no, date

def _extract_details(details_area, details_headers_text):
    words = details_area.extract_words()
    lines = __words_to_lines(words)
    
    dataCells = __group_line_items_as_cells(lines, SPACE_BETWEEN_WORDS)

    headers = [__get_cell_text(cell) for cell in dataCells[0]]
    data = [__get_cell_text(cell) for cell in dataCells[1]]

    result = dict(zip(headers, data))

    client = result.get('BILL TO', '')
    po_number = result.get('DETAILS', '').replace('PO Number:', '').strip()
    due_date = result.get('PAYMENT', '').replace('Due Date:', '').strip()

    return client, po_number, due_date

def _extract_table_data(table_area, headers_words):
    # get table lines
    textLines = __words_to_lines(table_area.extract_words())

    # group line items into cells
    tableData = __group_line_items_as_cells(textLines[1:])

    # make list of objects with header names as keys and data as values.
    table = []
    for td_row in tableData:
        elem = {}
        for idx, cell in enumerate(td_row):
            elem[headers_words[idx]['text']] = __get_cell_text(cell)
        table.append(elem)
    
    return table

def _extract_totals(line_area, docCenter):
    row = [[], []]
    for x in line_area.extract_words():
        idx = 0
        if x['x1'] > docCenter:
            idx = 1

        row[idx].append(x['text'])
    row = [' '.join(col) for col in row]

    return {row[0]: row[1]}


def extract_data(file_path: Path):
    try:
        with pdfplumber.open(file_path) as pdf:
            page = pdf.pages[0]
            words = page.extract_words()
            rects = page.rects

        r = rects[0]
        doc_header_left = r['x0']
        doc_header_top = 20
        doc_header_right = r['x1']
        doc_header_bottom = page.height - r['y1'] - 10

        details_text = ['BILL', 'PAYMENT', 'Due']
        details_headers_text = ['BILL TO', 'DETAILS', 'PAYMENT']
        table_headers_text = ['ITEM', 'QUANTITY', 'RATE', 'AMOUNT']
        totals_text = ['Subtotal', 'VAT', 'Total']

        details_words = []
        table_headers_words = []
        totals_words = []

        for w in words:
            if w['text'] in details_text:
                details_words.append(w)
            elif w['text'] in totals_text:
                totals_words.append(w)
                if len(totals_words) == len(totals_text):
                    break
            elif w['text'] in table_headers_text:
                table_headers_words.append(w)

        details_top = details_words[0]['top']
        details_left = details_words[0]['x0']
        details_right = details_words[1]['x1'] + 55
        details_bottom = details_words[-1]['bottom']


        table_top = table_headers_words[0]['top']
        table_left = table_headers_words[0]['x0'] 
        table_right = table_headers_words[-1]['x1']
        table_bottom = totals_words[0]['top'] - 30

        subtotal_top = totals_words[0]['top']
        subtotal_bottom = totals_words[0]['bottom']        

        vat_top = totals_words[1]['top']
        vat_bottom = totals_words[1]['bottom']        

        total_top = totals_words[-1]['top']
        total_bottom = totals_words[-1]['bottom']  


        doc_header_bbox = (doc_header_left, doc_header_top, doc_header_right, doc_header_bottom)  
        details_bbox = (details_left, details_top, details_right, details_bottom)    
        table_bbox = (table_left, table_top, table_right, table_bottom)
        subtotal_bbox = (table_left, subtotal_top, table_right, subtotal_bottom)
        vat_bbox = (table_left, vat_top, table_right, vat_bottom)
        total_bbox = (table_left, total_top, table_right, total_bottom)
        # draw_boxes(INPUT_FILE, 'debug.pdf', [doc_header_bbox, details_bbox, table_bbox, subtotal_bbox, vat_bbox, total_bbox])

        doc_header_area = page.crop(doc_header_bbox)
        details_area = page.crop(details_bbox)

        table_area = page.crop(table_bbox)
        subtotal_area = page.crop(subtotal_bbox)
        vat_area = page.crop(vat_bbox)
        total_area = page.crop(total_bbox)

        company_name, invoice_no, date = _extract_doc_header(doc_header_area)
        client_name, po_no, due_date = _extract_details(details_area, details_headers_text)


        table = _extract_table_data(table_area, table_headers_words)

        docCenter = page.width / 2
        subtotal = _extract_totals(subtotal_area, docCenter)
        vat = _extract_totals(vat_area, docCenter)
        total = _extract_totals(total_area, docCenter)

        summary = {}
        summary.update(subtotal)
        summary.update(vat)
        summary.update(total)


        return {
            "invoice_number": invoice_no,
            "invoice_date": date,
            "due_date": due_date,
            "po_number": po_no,

            "vendor": company_name,
            "client": client_name,

            "items": table,
            "summary": summary
        }

                
    except Exception as e:
        print(f'Exception: {e}')


def gen_excel(json_data, output_path): 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Invoice Report {json_data['invoice_date']}"
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name='Calibri Light', size=10)    

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15    

    ws['A1'] = 'INVOICE REPORT'
    ws['A1'].font = Font(size= 18, bold= True, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    ws.merge_cells('A1:H2')


    meta_headers = {
        'A4': 'Vendor',     'A5': json_data['vendor'],
        'B4': 'Client',     'B5': json_data['client'],
        'C4': 'Invoice #',  'C5': json_data['invoice_number'],
        'D4': 'Date',       'D5': json_data['invoice_date'],
        'E4': 'Due Date',   'E5': json_data['due_date'],
        'H4': 'PO Number',  'H5': json_data['po_number'],
    }
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)    
    for cell_ref, value in meta_headers.items():
        cell = ws[cell_ref]
        cell.value = value
        cell.fill = PatternFill(start_color="F0F4FA", end_color="F0F4FA", fill_type="solid")
        cell.border = border
        row = int(cell_ref[1:])
        if row == 4:
            cell.font = Font(bold=True, color='888888', size=8)
        else:
            cell.font = Font(bold=True, size=10)    


    items = json_data['items']
    headers = list(items[0].keys())

    for idx, header in enumerate(headers, start= 1):
        cell = ws.cell(row= 7, column= idx, value= header)
        cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        cell.font = Font(color="FFFFFF")

    thin_border = Border(
        bottom=Side(style='thin', color='E0E0E0')
    )    
    for rowIdx, key in enumerate(items, start= 8):
        bg_color = "F5F5F5" if rowIdx % 2 == 0 else "FFFFFF"
        for colIdx, value in enumerate(key.values(), start= 1):
            cell = ws.cell(row= rowIdx, column= colIdx, value= value)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.border = thin_border
            if colIdx > 1:
                cell.alignment = Alignment(horizontal="right")            


    last_item_row = rowIdx + 2
    summary = json_data['summary']
    summary_keys = list(summary.keys())

    for i, key in enumerate(summary_keys[:-1]): 
        r = last_item_row + i
        ws.cell(row=r, column=3, value=key).font = Font(color='888888')
        ws.cell(row=r, column=4, value=summary[key]).alignment = Alignment(horizontal="right")

    total_row = last_item_row + len(summary_keys) - 1
    for col in range(1, 5):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    ws.cell(row=total_row, column=3, value='Total')
    ws.cell(row=total_row, column=4, value=summary['Total']).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=4).font = Font(color="FFFFFF", bold=True)


    wb.save(output_path)


if __name__ == "__main__":
    base_dir = Path(__file__).parent

    data = extract_data(base_dir / INPUT_FILE)

    gen_excel(data, base_dir / OUTPUT_FILE)

    print("Invoice generated:", base_dir / OUTPUT_FILE)